import openpyxl
import  os

def open_workbook():
    ROOT_DIR = os.path.dirname(os.path.abspath(__file__))
    screenshotpath = os.path.join(os.path.sep, ROOT_DIR,)
    # print (screenshotpath)
    return screenshotpath
   


class HomePageData:
    # test_HomePage_Data=[{"name":"deepa","email":"deepa","Password":"Ram"},{"name":"sahasra","email":"ram","Password":"ravan"}]

    @staticmethod
    def getTestData(test_case_name):
        dict = {}
        screenshotpath=open_workbook()
        book = openpyxl.load_workbook(str(screenshotpath)+r"\test_data.xlsx",read_only = True)
        sheet = book.active
        
        for i in range(1,sheet.max_row + 1):
            if sheet.cell(row=i,column=1).value == test_case_name:
                for j in range(2,sheet.max_column + 1):
                    dict[sheet.cell(row=1,column=j).value] = sheet.cell(row=i,column=j).value

        return[dict]